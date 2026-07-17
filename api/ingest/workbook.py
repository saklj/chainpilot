"""Strict .xlsx loading shared by template preview and daily validation."""

from __future__ import annotations

from io import BytesIO
from zipfile import BadZipFile

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from ingest.errors import IngestError

# Sized for real-world exports (tens of thousands of rows is normal); the caps
# exist to bound server memory, not to describe typical files.
MAX_FILE_BYTES = 20 * 1024 * 1024
MAX_DATA_ROWS = 50_000


def read_first_sheet(file_bytes: bytes) -> tuple[list[str], list[tuple[object, ...]]]:
    if len(file_bytes) > MAX_FILE_BYTES:
        raise IngestError("file_too_large", "Excel 文件不能超过 20MB", status_code=413)
    if not file_bytes:
        raise IngestError("invalid_xlsx", "Excel 文件为空")
    try:
        workbook = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    except (BadZipFile, InvalidFileException, KeyError, OSError, ValueError) as error:
        raise IngestError("invalid_xlsx", "文件不是有效的 .xlsx 工作簿") from error

    try:
        sheet = workbook.worksheets[0]
        iterator = sheet.iter_rows(values_only=True)
        first_row = next(iterator, None)
        if first_row is None:
            raise IngestError("missing_header", "Excel 首个工作表没有表头")
        columns = [str(value).strip() if value is not None else "" for value in first_row]
        while columns and not columns[-1]:
            columns.pop()
        if not columns or any(not column for column in columns):
            raise IngestError("invalid_header", "表头列名不能为空")
        if len(set(columns)) != len(columns):
            raise IngestError("duplicate_header", "表头列名不能重复")

        rows: list[tuple[object, ...]] = []
        for row_number, row in enumerate(iterator, start=2):
            if row_number > MAX_DATA_ROWS + 1:
                raise IngestError(
                    "too_many_rows",
                    f"Excel 数据行不能超过 {MAX_DATA_ROWS} 行",
                    status_code=413,
                )
            values = tuple(
                row[index] if index < len(row) else None for index in range(len(columns))
            )
            rows.append(values)
        return columns, rows
    finally:
        workbook.close()
