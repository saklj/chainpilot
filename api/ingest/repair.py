"""Deterministic format-only repair rules for validated Excel uploads.

Each cleaning tool is one ``RepairRule`` in ``REPAIR_RULES``. Extending the
pipeline means registering another deterministic rule; the orchestration stays
unchanged. Rules never perform fuzzy matching, inference, randomness, or LLM
calls, and they are attempted only for fields already rejected by validation.
"""

from __future__ import annotations

import re
import unicodedata
from collections.abc import Callable
from dataclasses import dataclass
from datetime import date, datetime
from io import BytesIO
from typing import Any

import duckdb
from openpyxl import load_workbook

from ingest.database import load_template
from ingest.errors import IngestError
from ingest.models import ValidationReport
from ingest.pipeline import validate_file
from ingest.workbook import read_first_sheet


@dataclass(frozen=True)
class RepairRule:
    name: str
    field: str | tuple[str, ...]
    apply: Callable[[object], object | None]

    def applies_to(self, field: str) -> bool:
        return field == self.field if isinstance(self.field, str) else field in self.field


@dataclass(frozen=True)
class Repair:
    row: int
    field: str
    original_value: Any
    new_value: Any
    rule_name: str


@dataclass(frozen=True)
class RepairOutcome:
    repairs: tuple[Repair, ...]
    report: ValidationReport


def _date_format(value: object) -> str | None:
    if not isinstance(value, str):
        return None
    text = value.strip()
    formats: tuple[tuple[str, str], ...] = (
        (r"\d{4}/\d{1,2}/\d{1,2}", "%Y/%m/%d"),
        (r"\d{4}\.\d{1,2}\.\d{1,2}", "%Y.%m.%d"),
        (r"\d{4}年\d{1,2}月\d{1,2}日", "%Y年%m月%d日"),
        (r"\d{8}", "%Y%m%d"),
    )
    for pattern, date_format in formats:
        if re.fullmatch(pattern, text) is None:
            continue
        try:
            return datetime.strptime(text, date_format).date().isoformat()
        except ValueError:
            return None
    return None


def _qty_format(value: object) -> int | None:
    if not isinstance(value, str):
        return None
    text = unicodedata.normalize("NFKC", value).strip()
    if re.fullmatch(r"[1-9]\d{0,2}(?:,\d{3})+", text):
        text = text.replace(",", "")
    elif re.fullmatch(r"\d+\.0+", text):
        text = text.split(".", 1)[0]
    elif re.fullmatch(r"\d+", text) is None:
        return None
    parsed = int(text)
    return parsed if parsed > 0 else None


def _key_normalize(value: object) -> str | None:
    if not isinstance(value, str):
        return None
    normalized = unicodedata.normalize("NFKC", value).strip().upper()
    return normalized if normalized and normalized != value else None


REPAIR_RULES = (
    RepairRule("date_format", "eta_date", _date_format),
    RepairRule("qty_format", "qty", _qty_format),
    RepairRule(
        "key_normalize",
        ("po_id", "material_pn", "supplier_id"),
        _key_normalize,
    ),
)


def _key_candidate_is_valid(
    connection: duckdb.DuckDBPyConnection, field: str, candidate: object
) -> bool:
    if not isinstance(candidate, str) or not candidate:
        return False
    if field == "material_pn":
        return (
            connection.execute(
                "SELECT 1 FROM materials WHERE material_pn = ?", [candidate]
            ).fetchone()
            is not None
        )
    if field == "supplier_id":
        return (
            connection.execute(
                "SELECT 1 FROM suppliers WHERE supplier_id = ?", [candidate]
            ).fetchone()
            is not None
        )
    if field == "po_id":
        return (
            connection.execute("SELECT 1 FROM open_po WHERE po_id = ?", [candidate]).fetchone()
            is None
        )
    return True


def _json_value(value: object) -> Any:
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return value


def repair_file(
    connection: duckdb.DuckDBPyConnection,
    file_bytes: bytes,
    filename: str,
) -> RepairOutcome:
    """Repair rejected format cells, then revalidate the complete workbook."""
    initial_report = validate_file(connection, file_bytes, filename=filename)
    template = load_template(connection)
    if template is None:  # validate_file already raises this; retained for typing safety.
        raise IngestError("template_not_found", "请先注册 Excel 列映射模板", status_code=404)
    mapping, _ = template
    columns, raw_rows = read_first_sheet(file_bytes)
    column_indexes = {target: columns.index(source) for target, source in mapping.items()}
    mutable_rows = [list(row) for row in raw_rows]
    repairs: list[Repair] = []
    repaired_cells: set[tuple[int, str]] = set()

    for error in initial_report.errors:
        cell_key = (error.row, error.field)
        if cell_key in repaired_cells or error.field not in column_indexes:
            continue
        row_index = error.row - 2
        column_index = column_indexes[error.field]
        original_value = mutable_rows[row_index][column_index]
        for rule in REPAIR_RULES:
            if not rule.applies_to(error.field):
                continue
            candidate = rule.apply(original_value)
            if candidate is None:
                continue
            if rule.name == "key_normalize" and not _key_candidate_is_valid(
                connection, error.field, candidate
            ):
                continue
            mutable_rows[row_index][column_index] = candidate
            repaired_cells.add(cell_key)
            repairs.append(
                Repair(
                    row=error.row,
                    field=error.field,
                    original_value=_json_value(original_value),
                    new_value=_json_value(candidate),
                    rule_name=rule.name,
                )
            )
            break

    if not repairs:
        return RepairOutcome(repairs=(), report=initial_report)

    workbook = load_workbook(BytesIO(file_bytes))
    try:
        sheet = workbook.worksheets[0]
        for repair in repairs:
            source_column = mapping[repair.field]
            sheet.cell(
                row=repair.row,
                column=columns.index(source_column) + 1,
                value=repair.new_value,
            )
        output = BytesIO()
        workbook.save(output)
    finally:
        workbook.close()

    fresh_report = validate_file(connection, output.getvalue(), filename=filename)
    return RepairOutcome(repairs=tuple(repairs), report=fresh_report)
