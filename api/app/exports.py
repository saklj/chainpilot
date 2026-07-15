"""In-memory exports for structured weekly-report data."""

from __future__ import annotations

import unicodedata
from collections.abc import Iterable, Sequence
from datetime import date
from decimal import Decimal
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from agent.report import ReportData


def _number(value: Decimal | int) -> int | float:
    if isinstance(value, int):
        return value
    return int(value) if value == value.to_integral() else float(value)


def _display_width(value: Any) -> int:
    text = value.isoformat() if isinstance(value, date) else str(value if value is not None else "")
    return sum(
        2 if unicodedata.east_asian_width(character) in {"W", "F", "A"} else 1
        for character in text
    )


def _populate_sheet(
    worksheet: Worksheet,
    headers: Sequence[str],
    rows: Iterable[Sequence[Any]],
) -> None:
    worksheet.append(list(headers))
    for row in rows:
        worksheet.append(list(row))
    worksheet.freeze_panes = "A2"
    for cell in worksheet[1]:
        cell.font = Font(bold=True)

    for column_index, column_cells in enumerate(worksheet.iter_cols(), start=1):
        width = max(_display_width(cell.value) for cell in column_cells)
        worksheet.column_dimensions[get_column_letter(column_index)].width = min(
            max(width + 2, 10), 60
        )


def build_report_workbook(data: ReportData) -> bytes:
    """Build the four-table planner workbook without touching the filesystem."""
    workbook = Workbook()
    workbook.remove(workbook.active)

    kpi = workbook.create_sheet("KPI总览")
    _populate_sheet(
        kpi,
        ["指标", "数值"],
        [
            ["红色风险物料", data.kpi.red_count],
            ["橙色风险物料", data.kpi.orange_count],
            ["黄色风险物料", data.kpi.yellow_count],
            ["绿色风险物料", data.kpi.green_count],
            ["预计总缺口量", data.kpi.total_gap_qty],
            ["红橙物料占比", float(data.kpi.red_orange_pct / Decimal(100))],
            ["红色风险环比", data.comparison.red_change if data.comparison.red_change is not None else "—"],
            [
                "橙色风险环比",
                data.comparison.orange_change if data.comparison.orange_change is not None else "—",
            ],
        ],
    )
    for row in range(2, 7):
        kpi.cell(row=row, column=2).number_format = "#,##0"
    kpi.cell(row=7, column=2).number_format = "0.00%"
    for row in (8, 9):
        if isinstance(kpi.cell(row=row, column=2).value, int):
            kpi.cell(row=row, column=2).number_format = "+0;-0;0"

    top_risks = workbook.create_sheet("Top风险物料")
    _populate_sheet(
        top_risks,
        ["物料料号", "物料名称", "DOI（天）", "缺口量", "断料日", "风险原因"],
        (
            [
                item.material_pn,
                item.material_name,
                _number(item.doi_days),
                item.gap_qty,
                item.gap_date,
                item.risk_reasons,
            ]
            for item in data.top_risks
        ),
    )
    for row in range(2, top_risks.max_row + 1):
        top_risks.cell(row=row, column=3).number_format = "0.0"
        top_risks.cell(row=row, column=4).number_format = "#,##0"
        top_risks.cell(row=row, column=5).number_format = "yyyy-mm-dd"

    suppliers = workbook.create_sheet("供应商敞口")
    _populate_sheet(
        suppliers,
        ["供应商 ID", "供应商名称", "关联红橙物料数", "加权缺口量"],
        (
            [
                item.supplier_id,
                item.supplier_name,
                item.red_orange_material_count,
                _number(item.weighted_gap_qty),
            ]
            for item in data.supplier_exposure
        ),
    )
    for row in range(2, suppliers.max_row + 1):
        suppliers.cell(row=row, column=3).number_format = "#,##0"
        suppliers.cell(row=row, column=4).number_format = "#,##0.00"

    commodities = workbook.create_sheet("Commodity分布")
    _populate_sheet(
        commodities,
        ["Commodity", "红", "橙", "黄", "绿", "总缺口量"],
        (
            [
                item.commodity,
                item.red_count,
                item.orange_count,
                item.yellow_count,
                item.green_count,
                item.total_gap_qty,
            ]
            for item in data.commodity_distribution
        ),
    )
    for row in range(2, commodities.max_row + 1):
        for column in range(2, 7):
            commodities.cell(row=row, column=column).number_format = "#,##0"

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()
