"""Contract and real-data smoke tests for the FastAPI route layer."""

from __future__ import annotations

import json
from io import BytesIO
from pathlib import Path

import duckdb
import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from agent.llm import LLMResult, TokenUsage
from app.deps import get_llm
from app.main import app

REPO_ROOT = Path(__file__).resolve().parents[2]
REAL_DB = REPO_ROOT / "data" / "chainpilot.duckdb"
pytestmark = pytest.mark.skipif(not REAL_DB.exists(), reason="real DuckDB fixture is absent")


class SequenceLLM:
    """Return deterministic SQL and answer completions without network access."""

    def __init__(self, *responses: str) -> None:
        self.responses = list(responses)

    def chat(self, messages, *, temperature=0.0, timeout=30):
        del messages, temperature, timeout
        return LLMResult(
            content=self.responses.pop(0),
            usage=TokenUsage(prompt_tokens=8, completion_tokens=2),
        )


@pytest.fixture
def client() -> TestClient:
    with TestClient(app) as test_client:
        yield test_client
    app.dependency_overrides.clear()


def test_risk_summary_smoke(client: TestClient) -> None:
    response = client.get("/api/risk/summary")
    assert response.status_code == 200
    payload = response.json()
    assert payload["eval_date"]
    assert payload["red_count"] == 20
    assert payload["orange_count"] == 16
    assert payload["yellow_count"] == 66
    assert payload["green_count"] == 198
    assert payload["by_commodity"]
    assert len(payload["top_suppliers"]) == 5


def test_risk_material_filters_and_search(client: TestClient) -> None:
    red = client.get("/api/risk/materials", params={"level": "RED"})
    assert red.status_code == 200
    assert red.json()
    assert all(row["risk_level"] == "RED" for row in red.json())

    search = client.get("/api/risk/materials", params={"search": "PN-00003"})
    assert search.status_code == 200
    assert [row["material_pn"] for row in search.json()] == ["PN-00003"]


def test_risk_material_detail_and_404(client: TestClient) -> None:
    response = client.get("/api/risk/materials/PN-00003")
    assert response.status_code == 200
    payload = response.json()
    assert payload["material_pn"] == "PN-00003"
    assert "需求主要来自" in payload["explanation"]
    assert payload["top_skus"]
    assert payload["suppliers"]
    eval_date = client.get("/api/risk/summary").json()["eval_date"]
    assert all(row["eta_date"] > eval_date for row in payload["open_pos"])

    missing = client.get("/api/risk/materials/PN-NOT-FOUND")
    assert missing.status_code == 404
    assert missing.json()["detail"]["code"] == "material_not_found"


def test_forecast_endpoints_and_404(client: TestClient) -> None:
    sku_response = client.get("/api/forecast/skus")
    assert sku_response.status_code == 200
    skus = sku_response.json()
    assert skus == sorted(skus, key=lambda row: row["sku_id"])

    sku_id = skus[0]["sku_id"]
    forecast_response = client.get(f"/api/forecast/{sku_id}")
    assert forecast_response.status_code == 200
    payload = forecast_response.json()
    assert payload["sku_id"] == sku_id
    assert len(payload["history"]) == 90
    assert len(payload["forecast"]) == 84
    assert {row["model_name"] for row in payload["forecast"]} == {
        "seasonal_naive",
        "ets",
        "lightgbm",
    }

    missing = client.get("/api/forecast/NOT-A-SKU")
    assert missing.status_code == 404
    assert missing.json()["detail"]["code"] == "sku_not_found"


def test_forecast_metrics(client: TestClient) -> None:
    response = client.get("/api/forecast/metrics")
    assert response.status_code == 200
    payload = response.json()
    assert len(payload) == 9
    assert {row["model_name"] for row in payload} == {
        "seasonal_naive",
        "ets",
        "lightgbm",
    }
    assert all(
        len([row for row in payload if row["model_name"] == model]) == 3
        for model in {row["model_name"] for row in payload}
    )
    assert all(row["wmape"] > 0 for row in payload)
    assert payload == sorted(payload, key=lambda row: (row["model_name"], row["fold"]))


def test_chat_uses_injected_llm_and_serializes_verdict(client: TestClient) -> None:
    llm = SequenceLLM(
        "```sql\nSELECT gap_qty FROM material_risk "
        "WHERE material_pn = 'PN-00003' ORDER BY eval_date DESC LIMIT 1\n```",
        "缺口是13153。",
    )
    app.dependency_overrides[get_llm] = lambda: llm

    response = client.post("/api/chat", json={"question": "PN-00003的缺口是多少？"})
    assert response.status_code == 200
    payload = response.json()
    assert payload["rows"] == [[13153]]
    assert payload["row_count"] == 1
    assert payload["verdict"]["verdict"] == "pass"
    assert payload["verdict"]["matched"] == [{"value": "13153", "row": 0, "column": 0}]
    assert json.loads(json.dumps(payload, ensure_ascii=False)) == payload


def test_chat_empty_question_is_422(client: TestClient) -> None:
    app.dependency_overrides[get_llm] = lambda: SequenceLLM()
    response = client.post("/api/chat", json={"question": ""})
    assert response.status_code == 422


def _read_sse(response) -> list[dict]:
    events = []
    for line in response.iter_lines():
        if line.startswith("data: "):
            events.append(json.loads(line.removeprefix("data: ")))
    return events


def test_chat_stream_matches_sync_response(client: TestClient) -> None:
    responses = (
        "```sql\nSELECT gap_qty FROM material_risk "
        "WHERE material_pn = 'PN-00003' ORDER BY eval_date DESC LIMIT 1\n```",
        "缺口是13153。",
    )
    app.dependency_overrides[get_llm] = lambda: SequenceLLM(*responses)

    with client.stream(
        "POST", "/api/chat/stream", json={"question": "PN-00003的缺口是多少？"}
    ) as response:
        assert response.status_code == 200
        assert response.headers["content-type"].startswith("text/event-stream")
        events = _read_sse(response)

    assert [event["type"] for event in events] == [
        "stage",
        "sql",
        "rows",
        "answer",
        "result",
    ]
    assert events[1]["sql"].startswith("SELECT gap_qty")
    assert events[2]["rows"] == [[13153]]
    assert events[2]["row_count"] == 1

    app.dependency_overrides[get_llm] = lambda: SequenceLLM(*responses)
    sync = client.post("/api/chat", json={"question": "PN-00003的缺口是多少？"})
    assert sync.status_code == 200
    assert events[-1]["result"] == sync.json()


def test_chat_stream_no_answer_only_emits_stage_and_result(client: TestClient) -> None:
    app.dependency_overrides[get_llm] = lambda: SequenceLLM("NO_ANSWER")

    with client.stream(
        "POST", "/api/chat/stream", json={"question": "明天天气怎么样？"}
    ) as response:
        assert response.status_code == 200
        events = _read_sse(response)

    assert [event["type"] for event in events] == ["stage", "result"]
    assert events[-1]["result"]["refused"] is True
    assert events[-1]["result"]["refusal_reason"] == "out_of_scope"


def test_chat_stream_is_visible_in_openapi(client: TestClient) -> None:
    assert "/api/chat/stream" in client.get("/openapi.json").json()["paths"]


def test_report_endpoints_match_weekly_report(client: TestClient) -> None:
    latest = client.get("/api/report/latest")
    listing = client.get("/api/report/list")
    assert latest.status_code == 200
    assert listing.status_code == 200

    connection = duckdb.connect(str(REAL_DB), read_only=True)
    try:
        row = connection.execute(
            "SELECT report_date, content_md, narrative_fallbacks, created_at "
            "FROM weekly_report ORDER BY report_date DESC LIMIT 1"
        ).fetchone()
    finally:
        connection.close()
    payload = latest.json()
    assert payload == {
        "report_date": row[0].isoformat(),
        "content_md": row[1],
        "narrative_fallbacks": json.loads(row[2]),
        "created_at": row[3].isoformat(),
    }
    assert listing.json()[0]["report_date"] == payload["report_date"]


def test_report_by_date_matches_latest_and_returns_structured_404(
    client: TestClient,
) -> None:
    latest = client.get("/api/report/latest")
    report_date = latest.json()["report_date"]

    dated = client.get(f"/api/report/{report_date}")
    assert dated.status_code == 200
    assert dated.json() == latest.json()

    # 2015-01-01 predates all risk snapshots, so it stays absent even after
    # historical report backfills; nearby dates like 2016-05-15 may exist.
    missing = client.get("/api/report/2015-01-01")
    assert missing.status_code == 404
    assert missing.json()["detail"] == {
        "code": "report_not_found",
        "message": "Weekly report 2015-01-01 not found",
    }


def test_report_workbook_has_typed_four_sheet_export_and_404(
    client: TestClient,
) -> None:
    report_date = client.get("/api/report/latest").json()["report_date"]
    response = client.get(f"/api/report/{report_date}/xlsx")

    assert response.status_code == 200
    assert response.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert response.headers["content-disposition"] == (
        f"attachment; filename=chainpilot-weekly-{report_date}.xlsx"
    )
    workbook = load_workbook(BytesIO(response.content), data_only=False)
    assert workbook.sheetnames == ["KPI总览", "Top风险物料", "供应商敞口", "Commodity分布"]
    assert all(sheet.freeze_panes == "A2" for sheet in workbook.worksheets)
    red_count = workbook["KPI总览"]["B2"].value
    assert red_count == 20
    assert type(red_count) is int
    assert isinstance(workbook["供应商敞口"]["D2"].value, (int, float))

    missing = client.get("/api/report/2015-01-01/xlsx")
    assert missing.status_code == 404
    assert missing.json()["detail"]["code"] == "report_not_found"
